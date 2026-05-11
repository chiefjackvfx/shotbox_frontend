from __future__ import annotations

import re
import tempfile
from pathlib import Path
from urllib.parse import urljoin, urlparse

import requests


INVALID_SHEET_TITLE_CHARS = re.compile(r"[\[\]:*?/\\]")
MAX_SHEET_TITLE_LENGTH = 31
THUMBNAIL_ROW_HEIGHT = 54


def sanitize_excel_sheet_title(value: str, fallback: str = "Timeline") -> str:
    title = INVALID_SHEET_TITLE_CHARS.sub("_", str(value or "").strip())
    title = " ".join(title.split())
    if not title:
        title = fallback
    return title[:MAX_SHEET_TITLE_LENGTH]


def unique_excel_sheet_title(value: str, used_titles: set[str]) -> str:
    base = sanitize_excel_sheet_title(value)
    title = base
    suffix = 2
    while title.lower() in used_titles:
        suffix_text = f" {suffix}"
        title = f"{base[:MAX_SHEET_TITLE_LENGTH - len(suffix_text)]}{suffix_text}"
        suffix += 1
    used_titles.add(title.lower())
    return title


def iter_export_shots(timeline: dict, include_hidden: bool = False) -> list[dict]:
    shots = timeline.get("shots", []) if isinstance(timeline, dict) else []
    if include_hidden:
        return list(shots or [])
    return [shot for shot in shots or [] if not bool(shot.get("hidden"))]


def task_titles_summary(shot: dict, include_hidden: bool = False) -> str:
    tasks = shot.get("tasks", []) if isinstance(shot, dict) else []
    titles = []
    for task in tasks or []:
        if not include_hidden and bool(task.get("hidden")):
            continue
        title = str(task.get("title") or "").strip()
        if title:
            titles.append(title)
    return "\n".join(titles)


def default_export_filename(job_data: dict) -> str:
    title = str((job_data or {}).get("title") or "job").strip() or "job"
    safe_title = re.sub(r"[^A-Za-z0-9._-]+", "_", title).strip("._-") or "job"
    return f"{safe_title}_timeline_summary.xlsx"


def _thumbnail_url(value: str, base_url: str | None) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    parsed = urlparse(text)
    if parsed.scheme in {"http", "https"}:
        return text
    if text.startswith("/") and base_url:
        return urljoin(base_url.rstrip("/") + "/", text.lstrip("/"))
    return None


def _thumbnail_local_path(value: str) -> Path | None:
    text = str(value or "").strip()
    if not text:
        return None
    parsed = urlparse(text)
    if parsed.scheme:
        return None
    path = Path(text).expanduser()
    if path.is_file():
        return path
    return None


def resolve_thumbnail_for_excel(
    shot: dict,
    *,
    base_url: str | None = None,
    download_dir: Path | None = None,
    timeout: float = 5.0,
) -> Path | None:
    value = str((shot or {}).get("thumbnail") or "").strip()
    local_path = _thumbnail_local_path(value)
    if local_path is not None:
        return local_path

    if download_dir is None:
        return None

    url = _thumbnail_url(value, base_url)
    if not url:
        return None

    try:
        response = requests.get(url, timeout=timeout)
        response.raise_for_status()
    except Exception:
        return None

    suffix = Path(urlparse(url).path).suffix.lower()
    if suffix not in {".png", ".jpg", ".jpeg", ".gif", ".bmp"}:
        suffix = ".png"
    shot_id = str((shot or {}).get("id") or (shot or {}).get("title") or "thumb")
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", shot_id).strip("._-") or "thumb"
    path = download_dir / f"{safe_name}{suffix}"
    try:
        path.write_bytes(response.content)
    except Exception:
        return None
    return path if path.is_file() else None


def create_timeline_summary_workbook(
    job_data: dict,
    *,
    include_hidden: bool = False,
    base_url: str | None = None,
    thumbnail_download_dir: Path | None = None,
):
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as ExcelImage
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to export Excel files. Install shotbox_frontend requirements first."
        ) from exc

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    header_fill = PatternFill("solid", fgColor="2B2F36")
    header_font = Font(color="FFFFFF", bold=True)
    centered = Alignment(vertical="center")
    wrapped = Alignment(vertical="center", wrap_text=True)
    timelines = (job_data or {}).get("timelines", []) or []
    used_titles: set[str] = set()

    for timeline in timelines:
        title = unique_excel_sheet_title(timeline.get("title", "Timeline"), used_titles)
        worksheet = workbook.create_sheet(title)
        worksheet.freeze_panes = "A2"
        worksheet.column_dimensions["A"].width = 18
        worksheet.column_dimensions["B"].width = 32
        worksheet.column_dimensions["C"].width = 14
        worksheet.column_dimensions["D"].width = 42

        headers = ("Thumbnail", "Shot", "Duration", "Tasks")
        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = centered

        for row_index, shot in enumerate(iter_export_shots(timeline, include_hidden), start=2):
            worksheet.cell(row=row_index, column=2, value=str(shot.get("title") or ""))
            worksheet.cell(row=row_index, column=3, value=str(shot.get("duration") or ""))
            worksheet.cell(row=row_index, column=4, value=task_titles_summary(shot))
            worksheet.row_dimensions[row_index].height = THUMBNAIL_ROW_HEIGHT
            for column in range(1, 5):
                worksheet.cell(row=row_index, column=column).alignment = centered
            worksheet.cell(row=row_index, column=4).alignment = wrapped

            thumbnail_path = resolve_thumbnail_for_excel(
                shot,
                base_url=base_url,
                download_dir=thumbnail_download_dir,
            )
            if thumbnail_path is None:
                continue
            try:
                image = ExcelImage(str(thumbnail_path))
                image.width = 96
                image.height = 54
                # Excel stores images as drawing objects, not true cell values.
                # Use openpyxl's standard cell anchor for broad Excel/Sheets import
                # compatibility; manual two-cell anchors can make some importers reject
                # the workbook even though openpyxl can reload it.
                worksheet.add_image(image, f"A{row_index}")
            except Exception:
                continue

    if not workbook.worksheets:
        title = unique_excel_sheet_title("Timeline", used_titles)
        worksheet = workbook.create_sheet(title)
        worksheet.append(("Thumbnail", "Shot", "Duration", "Tasks"))

    return workbook


def export_job_timeline_summary(
    job_data: dict,
    destination: str | Path,
    *,
    include_hidden: bool = False,
    base_url: str | None = None,
) -> Path:
    output_path = Path(destination).expanduser()
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")
    with tempfile.TemporaryDirectory(prefix="shotbox_excel_thumbs_") as tmp:
        workbook = create_timeline_summary_workbook(
            job_data,
            include_hidden=include_hidden,
            base_url=base_url,
            thumbnail_download_dir=Path(tmp),
        )
        workbook.save(output_path)
    return output_path
